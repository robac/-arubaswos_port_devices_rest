import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

BACKGROUND_COLOR = "E1DFDF"

def export_excel(filename, ports):
    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet.column_dimensions['A'].width = 5
    sheet.column_dimensions['B'].width = 17
    sheet.column_dimensions['C'].width = 17
    sheet.column_dimensions['D'].width = 50

    crow = 1
    for p in ports.values():
        c = sheet.cell(row=crow, column=1)
        c.value = p.id
        if p.status == "OPER_UP":
            c.font = Font(bold=True, color='FFFF0000')
        else:
            c.font = Font(bold=True)
        c.fill = PatternFill(start_color=BACKGROUND_COLOR, end_color=BACKGROUND_COLOR, fill_type="solid")

        c = sheet.cell(row=crow, column=2)
        c.value = p.name
        c.fill = PatternFill(start_color=BACKGROUND_COLOR, end_color=BACKGROUND_COLOR, fill_type="solid")

        c = sheet.cell(row=crow, column=3)
        c.value = "{} MACs".format(len(p.macs))
        c.fill = PatternFill(start_color=BACKGROUND_COLOR, end_color=BACKGROUND_COLOR, fill_type="solid")

        c = sheet.cell(row=crow, column=4)
        c.fill = PatternFill(start_color=BACKGROUND_COLOR, end_color=BACKGROUND_COLOR, fill_type="solid")

        crow += 1
        for mac in p.macs:
            c = sheet.cell(row=crow, column=2)
            c.value = mac.mac
            c = sheet.cell(row=crow, column=3)
            c.value = mac.ip
            c = sheet.cell(row=crow, column=4)
            c.value = mac.vendor
            crow += 1

    wb.save(filename)